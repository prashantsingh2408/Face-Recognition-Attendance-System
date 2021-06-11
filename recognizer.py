#Incomplete
# Match face, detect name and save attendence.
import cv2
import os
from time import sleep
import pickle
import openpyxl
saved_create_excute_once = False
saved = []  # make track of saved student

def saved_create():
    wb = openpyxl.load_workbook("database.xlsx")
    # sheet = wb.get_sheet_by_name('Sheet')
    sheet = wb['Sheet']

    loop = 1
    while True:
        if (sheet.cell(row=loop, column=1).value == None):
            break;
        if (sheet.cell(row=loop, column=1).value not in saved):
            saved.append(sheet.cell(row=loop, column=1).value)
        # print("fetch_sheet", sheet.cell(row=loop, column=1).value)
        loop = loop + 1
def makeAttendance(name):
    saved_create()
    # get student input data
    data = name.split('_')
    
    #stop if data is already saved
    if (data[0] in saved):
        print('Your Attendance is taken,give chance to your friends')
        input("Press Enter to continue")
        return 0

    # make track of saved student
    saved.append(data[0])
    print('data', data)


    #find empty row in xl file to save 
    row_to_save = 1 # let empty row is 1 
    while True:
        wb = openpyxl.load_workbook("database.xlsx")
        sheet = wb.get_sheet_by_name('Sheet')
        if sheet.cell(row=row_to_save, column=1).value == None:
            print("Empty row found at", row_to_save)
            break
        else:
            print(row_to_save,"is not empty row")
            row_to_save = row_to_save + 1
            print("Incremented row_to_save = ",row_to_save)


    #Save in empty row
    row_to_save = str(row_to_save)
    sheet["A" + row_to_save] = data[0]
    sheet["B" + row_to_save] = data[1]
    sheet["C" + row_to_save] = data[2]
    sheet["D" + row_to_save] = 'present'
    wb.save(filename="database.xlsx")
    input("""One Student attendance is update in database.xlsx,\n
             Give Chance to you friend
             Press ENTER to continue""")

detector = cv2.CascadeClassifier("haarcascade_frontalface_default.xml")
recognizer = cv2.face.LBPHFaceRecognizer_create()
recognizer.read("tranner.yml")

labels ={}

with open("labels.pickle", 'rb') as f:
     orig_labels = pickle.load(f)
     labels = { k:v for v,k in orig_labels.items()}
     #orig_labes = {'name',id}
     #labels = {'id','name'}
     
#Take image
webcam = cv2.VideoCapture(0)  #start camera
time_of_recognition = 100 #Stop recognizing after some time
while True:
    check, frame = webcam.read() # start reading image
    gray = cv2.cvtColor(frame,cv2.COLOR_BGR2GRAY)
    faces = detector.detectMultiScale(gray,1.1,5)

    for (x,y,w,h) in faces:
        face_gray = gray[y:y+h,x:x+w] #roi->region of interest (faces in image)
        face_color = frame[y:y+h, x:x+w]

        try:
            id, conf = recognizer.predict(face_gray)
        except:
            print("Does not train yet,run tranner.py module,")
            input()#wait
        #predict the face 
        #conf conformation
        #0 means 100% confirm
        
        #print conformation no
        print(conf)

        if conf>=45 and conf<=85:
            print(labels[id])
            #Save attentance on xl file
            makeAttendance(labels[id])
        #Create rectangel
        color = (255,0,0)
        thickness = 2
        cv2.rectangle(frame, (x, y), (x + w, y + h), color, thickness)
    
    #Stop recognizing after some time
    print(time_of_recognition)
    time_of_recognition = time_of_recognition - 1
    if (time_of_recognition == 0):
        break;
    #END Stop recognizing after some time
    
    cv2.imshow("frame",frame)
    cv2.waitKey(1)
 


