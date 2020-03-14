
from tkinter import *
from tkinter import ttk
from openpyxl import *
#to install openpyxl: 
#open cmd(admin)
#command: pip install openpyxl

#read xl on sheet object
wb = load_workbook('record.xlsx')

#sheet object
sheet = wb.active 

def setDefultToXl():
    # resize the width of columns in 
	# excel spreadsheet 
	sheet.column_dimensions['A'].width = 30
	sheet.column_dimensions['B'].width = 10
	sheet.column_dimensions['C'].width = 10
	sheet.column_dimensions['D'].width = 20
	sheet.column_dimensions['E'].width = 20
	sheet.column_dimensions['F'].width = 40
	sheet.column_dimensions['G'].width = 50

	# write given data to an excel spreadsheet 
	# at particular location 
	sheet.cell(row=1, column=1).value = "Name"
	sheet.cell(row=1, column=2).value = "Course"
	sheet.cell(row=1, column=3).value = "Semester"
	sheet.cell(row=1, column=4).value = "Form Number"
	sheet.cell(row=1, column=5).value = "Contact Number"
	sheet.cell(row=1, column=6).value = "Email id"
	sheet.cell(row=1, column=7).value = "Address"

#def create_win(title = 'Registration form',geometry = '400x400',background = 'gray'):
#    global window 
#    window = Tk()
#    window.title(title)
#    window.geometry(geometry)
#    window.configure(background = background);
#Please corret the error in this approch.

window  = Tk()
window.title('Registration form')
window.geometry('400x400')
window.configure(background = 'gray')


def createLabel():
    Label(window ,text = "Name",width=22).grid(row = 0,column = 0)
    Label(window ,text = "Course",width=22).grid(row = 1,column = 0)
    Label(window,text="Stream",width=22).grid(row = 2, column = 0)
    Label(window ,text = "Email",width=22).grid(row = 3,column = 0)
    Label(window ,text = "Contact no",width=22).grid(row = 4,column = 0)
def createField():
    name_field   = Entry(window).grid(row = 0,column = 1)
    cource_field = Entry(window).grid(row = 1,column = 1)
    stream_field = Entry(window).grid(row = 2,column = 1)
    email_filed  = Entry(window).grid(row = 3,column = 1)
    contact_no   = Entry(window).grid(row = 4,column = 1)

def show():
    print("Thank you for submitting your response")

#what is this ?
def clicked():
   res = "Welcome to " + txt.get()
   lbl.configure(text= res)

createLabel()
createField()
btn = ttk.Button(window ,text="Submit",command=show).grid(row=5,column=0)
window.mainloop()




b